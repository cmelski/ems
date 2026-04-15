import os
from datetime import date
from functools import wraps
from flask import Flask, render_template, redirect, url_for, flash, Response, jsonify, request, send_from_directory
from dev.db import db_create
from dev.db.db_client import DBClient
import psycopg
import openpyxl
import pandas as pd
from sqlalchemy import create_engine, inspect
from flask_login import login_user, LoginManager, current_user, logout_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
import uuid
from openpyxl.drawing.image import Image
from openpyxl.styles import Font
from flask import send_file
import io
from io import BytesIO
import requests
import boto3

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads/expense-receipts")
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
DOWNLOAD_FOLDER = os.path.join(BASE_DIR, "downloads/expense-download")
app.config["DOWNLOAD_FOLDER"] = DOWNLOAD_FOLDER
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

s3 = boto3.client(
    "s3",
    aws_access_key_id=os.environ["AWS_ACCESS_KEY_ID"],
    aws_secret_access_key=os.environ["AWS_SECRET_ACCESS_KEY"],
    region_name=os.environ["AWS_REGION"]
)


def upload_to_s3(file):
    bucket = os.environ["S3_BUCKET"]

    filename = secure_filename(file.filename)
    unique_name = f"{uuid.uuid4()}_{filename}"

    s3.upload_fileobj(
        file,
        bucket,
        unique_name,
        ExtraArgs={"ContentType": file.content_type}
    )

    url = f"https://{bucket}.s3.{os.environ['AWS_REGION']}.amazonaws.com/{unique_name}"

    return unique_name, url


# Configure Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
# login_manager.login_view = "login"

db_create.create_db()
db_create.create_table()


@login_manager.user_loader
def load_user(user_id):
    # Open a new cursor/connection to your DB
    db_client = DBClient()
    # Fetch the user row by id
    row = db_client.get_user(user_id)
    # print(row)
    estate_user = db_client.check_estate_user(row[0])
    estate_id = estate_user[0]
    role = row[5]

    if row:
        # Reconstruct the same User object you passed to login_user()
        user = User(user_id=row[0], first_name=row[1], last_name=row[2],
                    email=row[3], password=row[4], estate=estate_id, role=role)
        # print(user.estate)
        return user
    else:
        return None  # Flask-Login will treat this as not logged in


login_manager = LoginManager()


class User:
    def __init__(self, user_id, first_name, last_name, email, password, estate, role, active=True):
        self.user_id = user_id
        self.first_name = first_name
        self.last_name = last_name
        self.email = email
        self.password = password
        self.estate = estate
        self.active = active
        self.role = role

    # Flask-Login required methods:
    @property
    def is_authenticated(self):
        return True

    @property
    def is_active(self):
        return self.active

    @property
    def is_anonymous(self):
        return False

    def get_id(self):
        # Must return a string
        return str(self.user_id)


def logged_in_only(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if current_user.is_authenticated:
            return f(*args, **kwargs)
        return redirect(url_for('login'))  # 👈 better than abort

    return decorated_function


def roles_required(*roles):
    def wrapper(fn):
        from functools import wraps

        @wraps(fn)
        def decorated(*args, **kwargs):
            if current_user.role not in roles:
                return jsonify({"error": "Unauthorized"}), 403
            return fn(*args, **kwargs)

        return decorated

    return wrapper


def download_db_data():
    db_client = DBClient()
    data = db_client.get_table_data()
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet
    for sheet_name, rows in data.items():
        ws = wb.create_sheet(title=sheet_name)
        if rows:

            # Write rows
            for row in rows:
                ws.append(row)

    wb.save('output.xlsx')


# download_db_data()

def download_prod_db_data():
    conn = psycopg.connect(os.environ.get("PROD_ENGINE"))
    cursor = conn.cursor()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    # 👉 Define tables you want

    tables = ['asset', 'task', 'expense', 'bill', 'contact', 'note', 'settings', 'activity',
              'users', 'estate_users']

    for table in tables:
        cursor.execute(f"SELECT * FROM {table}")
        rows = cursor.fetchall()
        cols = [desc[0] for desc in cursor.description]

        ws = wb.create_sheet(title=table)

        # Header row
        ws.append(cols)

        # Data rows
        for row in rows:
            ws.append(row)

    cursor.close()
    conn.close()

    wb.save('prod_output.xlsx')


# download_prod_db_data()


def upload_db_data():
    db_client = DBClient()
    db_client.reset_db()
    # dev
    engine = create_engine(os.environ.get('DEV_ENGINE'))
    # prod (external DB URL)
    # engine = create_engine(os.environ.get('PROD_ENGINE')

    xls = pd.ExcelFile("prod_output.xlsx")
    inspector = inspect(engine)

    for table in xls.sheet_names:

        df = pd.read_excel(xls, sheet_name=table)

        # Skip empty sheets first
        if df.dropna(how="all").empty:
            print(f"Sheet '{table}' is empty, skipping.")
            continue

        cols = [c["name"] for c in inspector.get_columns(table)]

        df.columns = cols

        df.to_sql(table, engine, if_exists="append", index=False)

        print(f"Inserted {len(df)} rows into {table}")


# upload_db_data()

@app.route('/register_request', methods=["POST"])
def register_request():
    db_client = DBClient()
    user_info = []

    # check if user email already exists and raise an error
    email = request.form.get('email', '')
    result = db_client.check_existing_user(email)
    if result:
        flash("Email already registered. Log in instead.")
        return redirect(url_for("login"))
    else:
        first_name = request.form.get('first-name', '')
        last_name = request.form.get('last-name', '')
        password = request.form.get('password', '')
        hash_password = generate_password_hash(password, method='pbkdf2:sha256', salt_length=8)
        completed = 'no'
        date_requested = date.today().isoformat()
        user_info.extend([first_name, last_name, email, hash_password, completed, date_requested])
        new_request = db_client.register_request(user_info)
        if new_request:
            flash("Registration request successfully sent. You will be notified when you can log in.")
            return redirect(url_for('login'))
        else:
            flash("Registration request not successful. Please try again")
            return redirect(url_for('register'))


@app.route('/login_user', methods=["POST"])
def login_user_app():
    db_client = DBClient()
    # Find user by email entered
    email = request.form.get('email', '')
    result = db_client.check_existing_user(email)

    if not result:
        flash("Invalid email")
        return redirect(url_for('login'))
    else:
        # Check stored password hash against entered password hashed.
        password = request.form.get('password', '')
        if check_password_hash(result[4], password):
            # check if user belongs to any estates
            user_id = result[0]
            estate_user = db_client.check_estate_user(user_id)
            if estate_user:
                # Log in and authenticate user

                estate_id = estate_user[0]
                user = User(user_id=result[0], first_name=result[1], last_name=result[2],
                            email=result[3], password=result[4], role=result[5], estate=estate_id)
                login_user(user)
                # print(user.estate)
                # print(user.role)

                return redirect(url_for('home'))
            else:
                flash("User is not associated with for any active estates")
                return redirect(url_for('login'))
        else:
            flash("Invalid password")
            return redirect(url_for('login'))


@app.route('/logout', methods=['POST'])
@logged_in_only
def logout():
    logout_user()
    return redirect(url_for('home'))


def get_tasks():
    db_client = DBClient()
    rows = db_client.get_tasks_from_db(current_user.estate)
    return [
        {
            "id": r[0],
            "description": r[1],
            "category": r[2],
            "due_date": r[3],
            "priority": r[4],
            "status": r[5].lower()
        }
        for r in rows
    ]


@app.route('/api/tasks', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_tasks():
    tasks = get_tasks()
    return jsonify({
        "message": "Tasks returned successfully",
        "tasks": tasks
    })


def get_task_by_description(description):
    db_client = DBClient()
    row = db_client.get_task_by_description_from_db(description)
    return [
        {
            "id": row[0],
            "description": row[1],
            "category": row[2],
            "due_date": row[3],
            "priority": row[4],
            "status": row[5].lower()
        }
    ]


@app.route('/api/task', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_task_by_description():
    description = request.args.get("description")

    task = get_task_by_description(description)
    return jsonify({
        "message": "Task returned successfully",
        "task": task
    })


@app.route('/api/tasks', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def add_task():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        task_details = []
        description = data['description']
        category = data['category']
        due_date = data['due_date']
        priority = data['priority']
        status = data['status'].lower()
        estate_id = current_user.estate

        task_details.extend([description, category, due_date, priority, status, estate_id])

        new_task = db_client.add_task_to_db(task_details)
        # print(f'new task: {list(new_task)}')

        return jsonify({"message": "Task added successfully",
                        "task": {
                            "id": new_task[0],
                            "description": new_task[1],
                            "category": new_task[2],
                            "due_date": new_task[3],
                            "priority": new_task[4],
                            "status": new_task[5].lower(),
                            "estate_id": estate_id
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/tasks/<int:task_id>', methods=['DELETE'])
@logged_in_only
@roles_required("admin", "editor")
def delete_task_by_task_id(task_id):
    # print(f'task_id: {task_id}')
    db_client = DBClient()
    try:
        db_client.delete_task_by_task_id(task_id)
        return jsonify({"message": "Task deleted successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/tasks/<int:task_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_task_status_by_task_id(task_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    task_status = data['status']
    db_client = DBClient()
    try:
        db_client.update_task_status_by_task_id(task_id, task_status, data)
        return jsonify({"message": "Task updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/tasks/row/<int:task_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_task_row(task_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    db_client = DBClient()
    try:
        db_client.update_task_row(task_id, data)
        return jsonify({"message": "Task row updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_bills():
    db_client = DBClient()
    rows = db_client.get_bills_from_db(current_user.estate)
    return [
        {
            "id": r[0],
            "description": r[1],
            "amount": r[2],
            "due_date": r[3],
            "bill_type": r[4],
            "status": r[5].lower()
        }
        for r in rows
    ]


@app.route('/api/bills', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_bills():
    bills = get_bills()
    return jsonify({
        "message": "Bills returned successfully",
        "bills": bills
    })


@app.route('/api/bills', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def add_bill():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        bill_details = []
        description = data['description']
        amount = data['amount']
        due_date = data['due_date']
        bill_type = data['bill_type']
        status = data['status'].lower()
        estate_id = current_user.estate

        bill_details.extend([description, amount, due_date, bill_type, status, estate_id])

        new_bill = db_client.add_bill_to_db(bill_details)
        # print(f'new bill: {list(bill_details)}')

        return jsonify({"message": "Bill added successfully",
                        "bill": {
                            "id": new_bill[0],
                            "description": new_bill[1],
                            "amount": new_bill[2],
                            "due_date": new_bill[3],
                            "type": new_bill[4],
                            "status": new_bill[5].lower(),
                            "estate_id": estate_id
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/bills/<int:bill_id>', methods=['DELETE'])
@logged_in_only
@roles_required("admin", "editor")
def delete_bill_by_bill_id(bill_id):
    # print(f'bill_id: {bill_id}')
    db_client = DBClient()
    try:
        db_client.delete_bill_by_bill_id(bill_id)
        return jsonify({"message": "Bill deleted successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/bills/<int:bill_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_bill_status_by_bill_id(bill_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    bill_status = data['status']
    db_client = DBClient()
    try:
        db_client.update_bill_status_by_bill_id(bill_id, bill_status, data)
        return jsonify({"message": "Bill updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/bills/row/<int:bill_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_bill_row(bill_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    db_client = DBClient()
    try:
        db_client.update_bill_row(bill_id, data)
        return jsonify({"message": "Bill row updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_expenses():
    db_client = DBClient()
    rows = db_client.get_expenses_from_db(current_user.estate)
    return [
        {
            "id": r[0],
            "description": r[1],
            "amount": r[2],
            "date_incurred": r[3],
            "category": r[4],
            "notes": r[5],
            "reimbursable": r[6],
            "status": r[7].lower(),
            "estate_id": r[8],
            "receipt_path": r[9]
        }
        for r in rows
    ]


@app.route('/api/expenses', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_expenses():
    expenses = get_expenses()
    # print(expenses)
    # print(f'expenses: {expenses}')
    return jsonify({
        "message": "Expenses returned successfully",
        "expenses": expenses
    })


@app.route('/api/download-expenses', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_expenses_for_download():
    try:
        expenses = get_expenses()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Expenses"

        ws.append([
            "ID", "DESCRIPTION", "AMOUNT", "DATE_INCURRED",
            "CATEGORY", "PAYEE", "REIMBURSABLE", "STATUS", "RECEIPT"
        ])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        ws.column_dimensions["I"].width = 25

        for row_num, expense in enumerate(expenses, start=2):

            if isinstance(expense, dict):
                exp_id = expense.get('id')
                desc = expense.get('description')
                amount = expense.get('amount')
                date = expense.get('date_incurred')
                category = expense.get('category')
                notes = expense.get('notes')
                reimb = expense.get('reimbursable')
                status = expense.get('status')
                receipt_url = expense.get('receipt_path')  # NOW FULL S3 URL
            else:
                exp_id = expense[0]
                desc = expense[1]
                amount = expense[2]
                date = expense[3]
                category = expense[4]
                notes = expense[5]
                reimb = expense[6]
                status = expense[7]
                receipt_url = expense[9] if len(expense) > 9 else None

            ws.append([
                exp_id, desc, amount, date,
                category, notes, reimb, status, ""
            ])

            # =========================
            # RECEIPT HANDLING (S3 ONLY)
            # =========================
            if receipt_url:
                sheet_name = f"Receipt - {exp_id}"[:31]
                receipt_ws = wb.create_sheet(title=sheet_name)

                receipt_ws["A1"] = f"Receipt for Expense {exp_id}"
                receipt_ws["A2"] = f"Description: {desc}"

                # Back link
                receipt_ws["A3"] = "Back to Expenses"
                receipt_ws["A3"].hyperlink = "#'Expenses'!A1"
                receipt_ws["A3"].style = "Hyperlink"

                # Direct link to S3
                receipt_ws["A4"] = "Open Receipt in Browser"
                receipt_ws["A4"].hyperlink = receipt_url
                receipt_ws["A4"].style = "Hyperlink"

                # # Optional: embed image directly from S3
                # try:
                #     response = requests.get(receipt_url, timeout=5)
                #     p# rint("RECEIPT URL:", receipt_url)
                #
                #     if response.status_code == 200:
                #         img = Image(BytesIO(response.content))
                #
                #         max_width = 600
                #         if img.width > max_width:
                #             ratio = max_width / img.width
                #             img.width = int(img.width * ratio)
                #             img.height = int(img.height * ratio)
                #
                #         receipt_ws.add_image(img, "A6")
                #
                #     else:
                #         receipt_ws["A6"] = "Image not accessible"
                #
                # except Exception as e:
                #     print("Image embed skipped:", e)
                #     receipt_ws["A6"] = "Image preview unavailable"

                # Main sheet link
                cell = ws.cell(row=row_num, column=9)
                cell.value = "View Receipt"
                cell.hyperlink = f"#'{sheet_name}'!A1"
                cell.style = "Hyperlink"

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="expense_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("DOWNLOAD ERROR:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/download-bills', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_bills_for_download():
    try:
        bills = get_bills()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bills"

        ws.append([
            "ID", "DESCRIPTION", "AMOUNT", "DUE_DATE",
            "BILL_TYPE", "STATUS"
        ])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        ws.column_dimensions["I"].width = 25

        for row_num, bill in enumerate(bills, start=2):

            if isinstance(bill, dict):
                bill_id = bill.get('id')
                desc = bill.get('description')
                amount = bill.get('amount')
                due_date = bill.get('due_date')
                bill_type = bill.get('bill_type')
                status = bill.get('status')

            else:
                bill_id = bill[0]
                desc = bill[1]
                amount = bill[2]
                due_date = bill[3]
                bill_type = bill[4]
                status = bill[5]

            ws.append([
                bill_id, desc, amount, due_date,
                bill_type, status, ""
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="bill_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("DOWNLOAD ERROR:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/download-assets', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_assets_for_download():
    try:
        assets = get_assets()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Assets"

        ws.append([
            "ID", "NAME", "TYPE", "VALUE",
            "NOTES", "STATUS"
        ])

        for cell in ws[1]:
            cell.font = Font(bold=True)

        ws.freeze_panes = "A2"
        ws.column_dimensions["I"].width = 25

        for row_num, asset in enumerate(assets, start=2):

            if isinstance(asset, dict):
                asset_id = asset.get('id')
                name = asset.get('name')
                asset_type = asset.get('type')
                value = asset.get('value')
                notes = asset.get('location')
                status = asset.get('status')

            else:
                asset_id = asset[0]
                name = asset[1]
                asset_type = asset[2]
                value = asset[3]
                notes = asset[5]
                status = asset[6]

            ws.append([
                asset_id, name, asset_type, value,
                notes, status, ""
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="asset_export.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("DOWNLOAD ERROR:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/download-summary', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def download_financial_summary():
    try:
        assets = get_assets()
        bills = get_bills()
        expenses = get_expenses()

        wb = openpyxl.Workbook()

        # =========================
        # SUMMARY SHEET (FIRST TAB)
        # =========================
        from openpyxl.styles import Font, Alignment
        from openpyxl.chart import BarChart, Reference
        from openpyxl.formatting.rule import CellIsRule

        summary_ws = wb.active
        summary_ws.title = "Summary"

        # --- CALCULATIONS ---
        total_assets = sum(a['value'] if isinstance(a, dict) else a[3] for a in assets)
        total_bills = sum(
            (b['amount'] if isinstance(b, dict) else b[2])
            for b in bills
            if (b['status'] if isinstance(b, dict) else b[5]) != 'paid'
        )
        total_expenses = sum(e['amount'] if isinstance(e, dict) else e[2] for e in expenses)

        net = total_assets - total_bills - total_expenses

        # --- TITLE ---
        summary_ws.merge_cells("A1:B1")
        summary_ws["A1"] = "Financial Summary"
        summary_ws["A1"].font = Font(size=16, bold=True)
        summary_ws["A1"].alignment = Alignment(horizontal="center")

        # --- HEADERS ---
        summary_ws["A3"] = "Metric"
        summary_ws["B3"] = "Amount"
        for cell in summary_ws["A3:B3"][0]:
            cell.font = Font(bold=True)

        # --- DATA ---
        summary_ws.append(["Total Assets", total_assets])
        summary_ws.append(["Outstanding Bills", total_bills])
        summary_ws.append(["Expenses", total_expenses])
        summary_ws.append(["Net Position", net])

        # --- FORMATTING ---
        summary_ws.column_dimensions["A"].width = 28
        summary_ws.column_dimensions["B"].width = 18

        for row in summary_ws.iter_rows(min_row=4, max_row=7, min_col=2):
            for cell in row:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")

        summary_ws.freeze_panes = "A4"

        # --- CONDITIONAL FORMATTING ---
        summary_ws.conditional_formatting.add(
            "B7",
            CellIsRule(operator="greaterThan", formula=["0"], font=Font(color="008000"))
        )
        summary_ws.conditional_formatting.add(
            "B7",
            CellIsRule(operator="lessThan", formula=["0"], font=Font(color="FF0000"))
        )

        # --- BAR CHART ---
        data = Reference(summary_ws, min_col=2, min_row=4, max_row=6)
        cats = Reference(summary_ws, min_col=1, min_row=4, max_row=6)

        chart = BarChart()
        chart.title = "Financial Overview"
        chart.add_data(data)
        chart.set_categories(cats)

        summary_ws.add_chart(chart, "D4")

        # =========================
        # ASSETS TAB
        # =========================
        assets_ws = wb.create_sheet(title="Assets")

        assets_ws.append(["ID", "NAME", "TYPE", "VALUE", "NOTES", "STATUS"])
        for cell in assets_ws[1]:
            cell.font = Font(bold=True)

        for asset in assets:
            if isinstance(asset, dict):
                row = [
                    asset.get('id'),
                    asset.get('name'),
                    asset.get('type'),
                    asset.get('value'),
                    asset.get('location'),
                    asset.get('status')
                ]
            else:
                row = [asset[0], asset[1], asset[2], asset[3], asset[5], asset[6]]

            assets_ws.append(row)

        # =========================
        # BILLS TAB
        # =========================
        bills_ws = wb.create_sheet(title="Bills")

        bills_ws.append(["ID", "DESCRIPTION", "AMOUNT", "DUE_DATE", "TYPE", "STATUS"])
        for cell in bills_ws[1]:
            cell.font = Font(bold=True)

        for bill in bills:
            if isinstance(bill, dict):
                row = [
                    bill.get('id'),
                    bill.get('description'),
                    bill.get('amount'),
                    bill.get('due_date'),
                    bill.get('bill_type'),
                    bill.get('status')
                ]
            else:
                row = [bill[0], bill[1], bill[2], bill[3], bill[4], bill[5]]

            bills_ws.append(row)

        # =========================
        # EXPENSES TAB (SIMPLIFIED)
        # =========================
        expenses_ws = wb.create_sheet(title="Expenses")

        expenses_ws.append([
            "ID", "DESCRIPTION", "AMOUNT", "DATE_INCURRED",
            "CATEGORY", "REIMBURSABLE", "STATUS", "PAYEE", "RECEIPT"
        ])

        for cell in expenses_ws[1]:
            cell.font = Font(bold=True)

        for expense in expenses:
            if isinstance(expense, dict):
                receipt = expense.get('receipt_path')
                row = [
                    expense.get('id'),
                    expense.get('description'),
                    expense.get('amount'),
                    expense.get('date_incurred'),
                    expense.get('category'),
                    expense.get('reimbursable'),
                    expense.get('status'),
                    expense.get('notes'),
                    "View Receipt" if receipt else ""
                ]
            else:
                receipt = expense[9] if len(expense) > 9 else None
                row = [
                    expense[0], expense[1], expense[2], expense[3],
                    expense[4], expense[6], expense[7], expense[5],
                    "View Receipt" if receipt else ""
                ]

            expenses_ws.append(row)

            if receipt:
                cell = expenses_ws.cell(row=expenses_ws.max_row, column=9)
                cell.hyperlink = receipt
                cell.style = "Hyperlink"

        # =========================
        # RETURN FILE
        # =========================
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name="financial_summary.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        print("SUMMARY DOWNLOAD ERROR:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def add_expense():
    db_client = DBClient()

    try:
        data = request.form

        description = data.get('description')
        amount = float(data.get('amount') or 0)
        date_incurred = data.get('date_incurred')
        category = data.get('category')
        notes = data.get('notes')
        reimbursable = data.get('reimbursable')

        status = 'n/a' if reimbursable == 'No' else (data.get('status') or 'unpaid').lower()

        estate_id = current_user.estate

        file = request.files.get("receipt")
        receipt_url = None

        if file and file.filename:
            _, receipt_url = upload_to_s3(file)

        expense_details = [
            description,
            amount,
            date_incurred,
            category,
            notes,
            reimbursable,
            status,
            estate_id,
            receipt_url
        ]

        new_expense = db_client.add_expense_to_db(expense_details)

        return jsonify({
            "message": "Expense added successfully",
            "expense": {
                "id": new_expense[0],
                "description": new_expense[1],
                "amount": new_expense[2],
                "date_incurred": new_expense[3],
                "category": new_expense[4],
                "notes": new_expense[5],
                "reimbursable": new_expense[6],
                "status": new_expense[7].lower(),
                "estate_id": estate_id,
                "receipt_path": receipt_url
            }
        }), 201

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/uploads/receipts/<filename>')
@logged_in_only
@roles_required("admin", "editor")
def uploaded_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename)


@app.route('/api/expenses/<int:expense_id>/receipt', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def upload_receipt(expense_id):
    file = request.files.get("receipt")

    if not file or not file.filename:
        return jsonify({"error": "No file provided"}), 400

    try:
        _, receipt_url = upload_to_s3(file)

        db_client = DBClient()
        db_client.update_receipt(expense_id, receipt_url)

        return jsonify({
            "message": "Receipt uploaded",
            "receipt_url": receipt_url
        }), 200

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses/<int:expense_id>', methods=['DELETE'])
@logged_in_only
@roles_required("admin", "editor")
def delete_expense_by_expense_id(expense_id):
    # print(f'expense_id: {expense_id}')
    db_client = DBClient()
    try:
        db_client.delete_expense_by_expense_id(expense_id)
        return jsonify({"message": "Expense deleted successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses/<int:expense_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_expense_status_by_expense_id(expense_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    expense_status = data['status']
    db_client = DBClient()
    try:
        db_client.update_expense_status_by_expense_id(expense_id, expense_status, data)
        return jsonify({"message": "Expense updated successfully"}), 200
    except Exception as e:
        # print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/expenses/row/<int:expense_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_expense_row(expense_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    db_client = DBClient()
    try:
        db_client.update_expense_row(expense_id, data)
        return jsonify({"message": "Expense row updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_assets():
    db_client = DBClient()
    rows = db_client.get_assets_from_db(current_user.estate)
    return [
        {
            "id": r[0],
            "name": r[1],
            "type": r[2],
            "value": r[3],
            "beneficiary": r[4],
            "location": r[5],
            "status": r[6].lower()
        }
        for r in rows
    ]


@app.route('/api/assets', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_assets():
    assets = get_assets()
    # print(f'assets: {assets}')
    return jsonify({
        "message": "Assets returned successfully",
        "assets": assets
    })


@app.route('/api/assets', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def add_asset():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        asset_details = []
        name = data['name']
        type = data['type']
        value = data['value']
        beneficiary = data['beneficiary']
        location = data['location']
        status = data['status'].lower()
        estate_id = current_user.estate

        asset_details.extend([name, type, value, beneficiary, location,
                              status, estate_id])

        new_asset = db_client.add_asset_to_db(asset_details)
        # print(f'new asset: {list(asset_details)}')

        return jsonify({"message": "Asset added successfully",
                        "asset": {
                            "id": new_asset[0],
                            "name": new_asset[1],
                            "type": new_asset[2],
                            "value": new_asset[3],
                            "benificiary": new_asset[4],
                            "location": new_asset[5],
                            "status": new_asset[6].lower(),
                            "estate_id": estate_id
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/assets/<int:asset_id>', methods=['DELETE'])
@logged_in_only
@roles_required("admin", "editor")
def delete_asset_by_asset_id(asset_id):
    # print(f'asset_id: {asset_id}')
    db_client = DBClient()
    try:
        db_client.delete_asset_by_asset_id(asset_id)
        return jsonify({"message": "Asset deleted successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/assets/<int:asset_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_asset_status_by_asset_id(asset_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    asset_status = data['status']
    db_client = DBClient()
    try:
        db_client.update_asset_status_by_asset_id(asset_id, asset_status, data)
        return jsonify({"message": "Asset updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/assets/row/<int:asset_id>', methods=['PATCH'])
@logged_in_only
@roles_required("admin", "editor")
def update_asset_row(asset_id):
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    db_client = DBClient()
    try:
        db_client.update_asset_row(asset_id, data)
        return jsonify({"message": "Asset row updated successfully"}), 200
    except Exception as e:
        # print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_contacts():
    db_client = DBClient()
    rows = db_client.get_contacts_from_db(current_user.estate)
    return [
        {
            "id": r[0],
            "name": r[1],
            "role": r[2],
            "phone": r[3],
            "email": r[4]
        }
        for r in rows
    ]


@app.route('/api/contacts', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor")
def fetch_contacts():
    contacts = get_contacts()
    return jsonify({
        "message": "Contacts returned successfully",
        "contacts": contacts
    })


@app.route('/api/contacts', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def add_contacts():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        contact_details = []
        name = data['name']
        role = data['role']
        phone = data['phone']
        email = data['email']
        estate_id = current_user.estate

        contact_details.extend([name, role, phone, email, estate_id])

        new_contact = db_client.add_contact_to_db(contact_details)
        # print(f'new contact: {list(contact_details)}')

        return jsonify({"message": "Contact added successfully",
                        "contact": {
                            "id": new_contact[0],
                            "name": new_contact[1],
                            "role": new_contact[2],
                            "phone": new_contact[3],
                            "email": new_contact[4],
                            "estate_id": estate_id
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/contacts/<int:contact_id>', methods=['DELETE'])
@logged_in_only
@roles_required("admin", "editor")
def delete_contact_by_contact_id(contact_id):
    # print(f'contact_id: {contact_id}')
    db_client = DBClient()
    try:
        db_client.delete_contact_by_contact_id(contact_id)
        return jsonify({"message": "Contact deleted successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_notes():
    db_client = DBClient()
    rows = db_client.get_notes_from_db(current_user.estate)
    return [
        {
            "id": r[0],
            "date": r[1],
            "title": r[2],
            "category": r[3],
            "content": r[4]
        }
        for r in rows
    ]


@app.route('/api/notes', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_notes():
    notes = get_notes()
    return jsonify({
        "message": "Notes returned successfully",
        "notes": notes
    })


@app.route('/api/notes', methods=['POST'])
@logged_in_only
@roles_required("admin", "editor")
def add_notes():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        note_details = []
        date = data['date']
        title = data['title']
        category = data['category']
        content = data['content']
        estate_id = current_user.estate

        note_details.extend([date, title, category, content, estate_id])

        new_note = db_client.add_note_to_db(note_details)
        # print(f'new note: {list(note_details)}')

        return jsonify({"message": "Note added successfully",
                        "note": {
                            "id": new_note[0],
                            "date": new_note[1],
                            "title": new_note[2],
                            "category": new_note[3],
                            "content": new_note[4],
                            "estate_id": estate_id
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/notes/<int:note_id>', methods=['DELETE'])
@logged_in_only
@roles_required("admin", "editor")
def delete_note_by_note_id(note_id):
    # print(f'note_id: {note_id}')
    db_client = DBClient()
    try:
        db_client.delete_note_by_note_id(note_id)
        return jsonify({"message": "Note deleted successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_settings():
    db_client = DBClient()
    estate_id = current_user.estate
    rows = db_client.get_settings_from_db(estate_id)
    return {
        "id": rows[0],
        "name": rows[1],
        "dod": rows[2],
        "executor": rows[3],
        "ref": rows[4]
    }


@app.route('/api/settings', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_settings():
    settings = get_settings()
    # print(settings)
    return jsonify({
        "message": "Settings returned successfully",
        "settings": settings
    })


@app.route('/api/settings', methods=['POST'])
@logged_in_only
@roles_required("admin")
def add_settings():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        settings_details = []
        name = data['name']
        dod = data['dod']
        executor = data['executor']
        ref = data['ref']

        settings_details.extend([name, dod, executor, ref])

        new_settings = db_client.add_settings_to_db(settings_details)
        # print(f'new settings: {list(settings_details)}')

        return jsonify({"message": "Settings added successfully",
                        "settingsDetails": {
                            "id": new_settings[0],
                            "name": new_settings[1],
                            "dod": new_settings[2],
                            "executor": new_settings[3],
                            "ref": new_settings[4]
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/api/settings/', methods=['PATCH'])
@logged_in_only
@roles_required("admin")
def update_settings():
    # print("Raw request data:", request.data)
    data = request.get_json(force=True)
    # print("Parsed data:", data)

    if not data:
        return jsonify({"error": "No JSON received"}), 400

    db_client = DBClient()
    try:
        db_client.update_settings(data)
        return jsonify({"message": "Settings updated successfully"}), 200
    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


def get_activity():
    db_client = DBClient()
    rows = db_client.get_activity_log(current_user.estate)
    return [
        {
            "activity_id": r[0],
            "date": r[1],
            "category": r[2],
            "description": r[3],
            "detail": r[4],
            "status": r[5].lower(),
            "note": r[6]
        }
        for r in rows
    ]


@app.route('/api/activity', methods=['GET'])
@logged_in_only
@roles_required("admin", "editor", "viewer")
def fetch_activity():
    activity = get_activity()
    # print(f'activity loaded: {activity}')
    return jsonify({
        "message": "Activity returned successfully",
        "activity": activity
    })


def get_registration_requests():
    db_client = DBClient()
    rows = db_client.get_registration_requests_from_db()
    return [
        {
            "user_id": r[0],
            "first_name": r[1],
            "last_name": r[2],
            "email": r[3],
            "password": r[4],
            "completed": r[5],
            "date_requested": r[6]
        }
        for r in rows
    ]


@app.route('/api/registration-requests', methods=['GET'])
@logged_in_only
@roles_required("admin")
def fetch_registration_requests():
    registration_requests = get_registration_requests()
    return jsonify({
        "message": "Activity returned successfully",
        "registration_requests": registration_requests
    })


@app.route('/api/process-registration-request', methods=['POST'])
@logged_in_only
@roles_required("admin")
def process_reg_request():
    db_client = DBClient()
    try:
        # print("Raw request data:", request.data)
        data = request.get_json(force=True)
        # print("Parsed data:", data)

        if not data:
            return jsonify({"error": "No JSON received"}), 400

        user_details = []
        user_id = data['user_id']
        first_name = data['first_name']
        last_name = data['last_name']
        email = data['email']
        password = data['password']
        role = data['role']

        user_details.extend([user_id, first_name, last_name, email, password, role])

        new_user = db_client.add_user(user_details)

        # insert to estate_users
        estate_id = current_user.estate
        print(estate_id)
        db_client.add_user_estate(user_id, estate_id)

        # update register_requests record (completed = 'yes')
        db_client.update_registration_request(user_id)

        return jsonify({"message": "Registration request processed successfully",
                        "userDetails": {
                            "user_id": new_user[0],
                            "first_name": new_user[1],
                            "last_name": new_user[2],
                            "email": new_user[3],
                            "password": new_user[4],
                            "role": new_user[5]
                        }
                        }), 201

    except Exception as e:
        print("Error:", str(e))
        return jsonify({"error": str(e)}), 500


@app.route('/')
@logged_in_only
@roles_required("admin", "editor", "viewer")
def home():
    return render_template("index.html", current_user=current_user)


@app.route("/register")
def register():
    return render_template("register.html")


@app.route("/login")
def login():
    return render_template("login.html")


if __name__ == "__main__":
    # app.run(debug=app.config.get("DEBUG", False), port=5002)
    app.run(host="0.0.0.0", port=5002, debug=app.config.get("DEBUG", False))
